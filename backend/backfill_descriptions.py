#!/usr/bin/env python3
"""
Backfill rate_card_description from the Excel template.
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook

# Set DATABASE_URL to Railway production
os.environ['DATABASE_URL'] = 'postgresql://postgres:fxkNRkTWAyPMATyFUthxrCEaNxLNKgcQ@postgres.railway.internal:5432/railway'

sys.path.insert(0, str(Path(__file__).parent))

from database import SessionLocal
from models import RateCardItem

def backfill():
    db = SessionLocal()
    
    # Load template
    backend_dir = Path(__file__).parent
    template_path = backend_dir / "reference_files" / "WestPark_FireDoor_CostSheet_v3_AlphaSights.xlsx"
    
    if not template_path.exists():
        print(f"❌ Template not found: {template_path}")
        return
    
    print(f"📂 Loading template: {template_path}")
    wb = load_workbook(template_path, data_only=True)
    
    # Extract B-code descriptions from Rate Card sheet (rows 22-33, columns A-C)
    b_code_descriptions = {}
    b_code_prices = {}
    
    if "Rate Card" in wb.sheetnames:
        ws = wb["Rate Card"]
        for row in range(22, 34):  # Rows 22-33
            code = ws.cell(row, 1).value  # Column A
            desc = ws.cell(row, 2).value  # Column B
            unit_val = ws.cell(row, 3).value  # Column C (unit price)
            qty_val = ws.cell(row, 4).value  # Column D (qty)
            
            if code:
                code = str(code).strip()
                if desc:
                    b_code_descriptions[code] = str(desc).strip()
                    print(f"  {code}: {desc}")
                
                # Calculate total price
                if unit_val and qty_val:
                    try:
                        unit_price = float(unit_val)
                        qty = float(qty_val)
                        total = unit_price * qty
                        b_code_prices[code] = f"£{total:.2f}"
                    except:
                        pass
    
    print(f"\n💾 Extracted {len(b_code_descriptions)} B-code descriptions")
    print(f"💷 Extracted {len(b_code_prices)} B-code prices")
    
    # Update database
    updated = 0
    rate_items = db.query(RateCardItem).all()
    
    for item in rate_items:
        # Split rate_card_code by "/" to handle multi-code mappings
        codes = [c.strip() for c in item.rate_card_code.split('/') if c.strip()]
        
        # Use first matching B-code
        for code in codes:
            if code in b_code_descriptions:
                item.rate_card_description = b_code_descriptions[code][:500]
                if not item.unit_price and code in b_code_prices:
                    item.unit_price = b_code_prices[code]
                updated += 1
                print(f"✓ Updated {item.art_code} -> {code}: {item.rate_card_description[:60]}...")
                break
    
    db.commit()
    db.close()
    
    print(f"\n✅ Updated {updated} rate card items")

if __name__ == "__main__":
    backfill()
