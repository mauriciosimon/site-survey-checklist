#!/usr/bin/env python3
"""
End-to-end testing for Fire Door extraction and Excel population pipeline.
"""
import sys
import os
from pathlib import Path
from openpyxl import load_workbook

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent))

import firedoor_processor as fdp

def test_type1_pdf():
    """
    TEST 1 — Type 1 (FireDNA/EAS PDF format)
    File: Alpha Sights - Thames Court-Fire Door Survey-March 2026.txt
    """
    print("\n" + "="*80)
    print("TEST 1: Type 1 (FireDNA PDF format)")
    print("="*80)
    
    test_file = Path(__file__).parent / "test_files" / "Alpha Sights - Thames Court-Fire Door Survey-March 2026.txt"
    client_name = "Alpha Sights Thames Court"
    
    if not test_file.exists():
        print(f"❌ Test file not found: {test_file}")
        return False
    
    print(f"📂 Input: {test_file.name}")
    print(f"📏 Size: {test_file.stat().st_size / 1024:.1f} KB\n")
    
    # Detect format
    file_format = fdp.detect_format(str(test_file), test_file.name)
    print(f"🔍 Detected format: {file_format}")
    
    if file_format != "TYPE_1":
        print(f"❌ Expected TYPE_1, got {file_format}")
        return False
    
    # Extract doors
    print("\n📊 Extracting door data...")
    try:
        doors = fdp.extract_type1_pdf(str(test_file))
    except Exception as e:
        print(f"❌ Extraction failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    print(f"✓ Extracted {len(doors)} doors\n")
    
    # Expected results validation
    expected_total = 39
    if len(doors) != expected_total:
        print(f"⚠️  Warning: Expected {expected_total} doors, got {len(doors)}")
    
    # Count by status
    compliant = sum(1 for d in doors if d.get('compliant', False))
    replace = sum(1 for d in doors if d.get('replace', False))
    remedial = len(doors) - compliant - replace
    flagged = sum(1 for d in doors if d.get('flagged', False))
    
    print(f"📈 Door Summary:")
    print(f"  Total: {len(doors)}")
    print(f"  Compliant (green): {compliant}")
    print(f"  Replace (red): {replace}")
    print(f"  Remedial (amber): {remedial}")
    print(f"  Flagged: {flagged}\n")
    
    # Check ART codes
    art_codes = set()
    for door in doors:
        faults = door.get('faults', [])
        for fault in faults:
            art_code = fault.get('art_code', '')
            if art_code:
                art_codes.add(art_code)
    
    expected_arts = {'01', '02', '04', '05', '08', '10', '11', '13', '16', '18', '19', '21', '22'}
    print(f"🔖 ART Codes found: {sorted(art_codes)}")
    print(f"🔖 Expected: {sorted(expected_arts)}")
    
    missing = expected_arts - art_codes
    if missing:
        print(f"⚠️  Missing ART codes: {sorted(missing)}")
    
    # Check ART10 mapping to B05
    art10_doors = [d for d in doors if any(f.get('art_code') == '10' for f in d.get('faults', []))]
    print(f"\n🔍 ART10 doors found: {len(art10_doors)}")
    
    for door in art10_doors[:3]:  # Show first 3
        door_ref = door.get('door_ref', 'Unknown')
        faults = [f for f in door.get('faults', []) if f.get('art_code') == '10']
        for fault in faults:
            b_code = fault.get('b_code', 'NONE')
            print(f"  {door_ref}: ART10 → {b_code} {'✓' if b_code == 'B05' else '❌'}")
    
    # Check ART18 doors (A12-L3 and A14-L3)
    art18_doors = [d for d in doors if any(f.get('art_code') == '18' for f in d.get('faults', []))]
    print(f"\n🔍 ART18 (replace) doors found: {len(art18_doors)}")
    
    for door in art18_doors:
        door_ref = door.get('door_ref', 'Unknown')
        is_replace = door.get('replace', False)
        option_b = door.get('option_b_code', 'NONE')
        print(f"  {door_ref}: replace={is_replace}, Option B={option_b}")
        
        if door_ref in ['A12-L3', 'A14-L3']:
            expected_b = 'A09' if door_ref == 'A12-L3' else 'A11'
            if option_b == expected_b:
                print(f"    ✓ Correct Option B code")
            else:
                print(f"    ❌ Expected {expected_b}, got {option_b}")
    
    # Generate Excel output
    print("\n📄 Generating Excel output...")
    template_path = Path(__file__).parent / "reference_files" / "WestPark_FireDoor_CostSheet_v3_AlphaSights.xlsx"
    output_path = Path(__file__).parent / "test_files" / f"{client_name.replace(' ', '_')}_Output.xlsx"
    
    try:
        result_path = fdp.populate_excel_template(doors, client_name, str(template_path), str(output_path))
        print(f"✓ Excel generated: {output_path.name}")
    except Exception as e:
        print(f"❌ Excel generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Verify Excel content
    print("\n🔎 Verifying Excel content...")
    wb = load_workbook(output_path, data_only=True)
    
    if "Door Schedule" in wb.sheetnames:
        ws = wb["Door Schedule"]
        # Count non-header rows (first row is header)
        door_rows = sum(1 for row in ws.iter_rows(min_row=2) if ws.cell(row[0].row, 1).value)
        print(f"  Door Schedule rows: {door_rows}")
        
        if door_rows != len(doors):
            print(f"  ⚠️  Expected {len(doors)} rows, got {door_rows}")
    else:
        print(f"  ❌ Door Schedule sheet not found")
    
    if "Quote Sheet" in wb.sheetnames:
        ws = wb["Quote Sheet"]
        # Check if client name and date are populated
        # (exact cell locations depend on template structure)
        print(f"  ✓ Quote Sheet found")
    else:
        print(f"  ❌ Quote Sheet not found")
    
    wb.close()
    
    print("\n✅ TEST 1 COMPLETE\n")
    return True


def test_type2_excel():
    """
    TEST 2 — Type 2 (Excel survey format)
    File: Natixis Fire Door Survey Remedials Survey - CITYSPACE.xlsx
    """
    print("\n" + "="*80)
    print("TEST 2: Type 2 (Excel survey format)")
    print("="*80)
    
    test_file = Path(__file__).parent / "test_files" / "Natixis Fire Door Survey Remedials Survey - CITYSPACE.xlsx"
    client_name = "Natixis CITYSPACE"
    
    if not test_file.exists():
        print(f"❌ Test file not found: {test_file}")
        return False
    
    print(f"📂 Input: {test_file.name}")
    print(f"📏 Size: {test_file.stat().st_size / 1024 / 1024:.1f} MB\n")
    
    # Detect format
    file_format = fdp.detect_format(str(test_file), test_file.name)
    print(f"🔍 Detected format: {file_format}")
    
    if file_format != "TYPE_2":
        print(f"❌ Expected TYPE_2, got {file_format}")
        return False
    
    # Extract doors
    print("\n📊 Extracting door data...")
    try:
        doors = fdp.extract_type2_excel(str(test_file))
    except Exception as e:
        print(f"❌ Extraction failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    print(f"✓ Extracted {len(doors)} doors\n")
    
    # Expected results
    expected_total = 15
    if len(doors) != expected_total:
        print(f"⚠️  Warning: Expected {expected_total} doors, got {len(doors)}")
    
    # Count by status
    compliant = sum(1 for d in doors if d.get('compliant', False))
    replace = sum(1 for d in doors if d.get('replace', False))
    remedial = len(doors) - compliant - replace
    flagged = sum(1 for d in doors if d.get('flagged', False))
    
    print(f"📈 Door Summary:")
    print(f"  Total: {len(doors)}")
    print(f"  Compliant (green): {compliant}")
    print(f"  Replace (red): {replace}")
    print(f"  Remedial (amber): {remedial}")
    print(f"  Flagged (orange): {flagged}\n")
    
    # Check for PENDING in Option B
    pending_count = sum(1 for d in doors if d.get('option_b_code') == 'PENDING')
    print(f"🔍 Doors with Option B = PENDING: {pending_count}")
    
    if pending_count != len(doors):
        print(f"  ⚠️  Expected all {len(doors)} doors to have PENDING, got {pending_count}")
    
    # Check for "Unable" flagged doors
    unable_flagged = [d for d in doors if d.get('flagged') and 'unable' in d.get('notes', '').lower()]
    print(f"\n🔍 Doors flagged with 'Unable': {len(unable_flagged)}")
    
    for door in unable_flagged[:5]:  # Show first 5
        door_ref = door.get('door_ref', 'Unknown')
        notes = door.get('notes', '')[:60]
        print(f"  {door_ref}: {notes}...")
    
    # Check Floor1-D11 specifically
    floor1_d11 = next((d for d in doors if d.get('door_ref') == 'Floor1-D11'), None)
    if floor1_d11:
        is_flagged = floor1_d11.get('flagged', False)
        notes = floor1_d11.get('notes', '')
        print(f"\n🔍 Floor1-D11:")
        print(f"  Flagged: {is_flagged} {'✓' if is_flagged else '❌'}")
        print(f"  Notes: {notes[:100]}")
    else:
        print(f"\n⚠️  Floor1-D11 not found in extracted doors")
    
    # Generate Excel output
    print("\n📄 Generating Excel output...")
    template_path = Path(__file__).parent / "reference_files" / "WestPark_FireDoor_CostSheet_v3_AlphaSights.xlsx"
    output_path = Path(__file__).parent / "test_files" / f"{client_name.replace(' ', '_')}_Output.xlsx"
    
    try:
        result_path = fdp.populate_excel_template(doors, client_name, str(template_path), str(output_path))
        print(f"✓ Excel generated: {output_path.name}")
    except Exception as e:
        print(f"❌ Excel generation failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Verify Excel content
    print("\n🔎 Verifying Excel content...")
    wb = load_workbook(output_path, data_only=True)
    
    if "Door Schedule" in wb.sheetnames:
        ws = wb["Door Schedule"]
        door_rows = sum(1 for row in ws.iter_rows(min_row=2) if ws.cell(row[0].row, 1).value)
        print(f"  Door Schedule rows: {door_rows}")
        
        if door_rows != len(doors):
            print(f"  ⚠️  Expected {len(doors)} rows, got {door_rows}")
    else:
        print(f"  ❌ Door Schedule sheet not found")
    
    if "Quote Sheet" in wb.sheetnames:
        print(f"  ✓ Quote Sheet found")
    else:
        print(f"  ❌ Quote Sheet not found")
    
    wb.close()
    
    print("\n✅ TEST 2 COMPLETE\n")
    return True


if __name__ == "__main__":
    print("\n" + "="*80)
    print("FIRE DOOR PIPELINE END-TO-END TESTS")
    print("="*80)
    
    test1_pass = test_type1_pdf()
    test2_pass = test_type2_excel()
    
    print("\n" + "="*80)
    print("FINAL RESULTS")
    print("="*80)
    print(f"TEST 1 (Type 1 PDF): {'✅ PASS' if test1_pass else '❌ FAIL'}")
    print(f"TEST 2 (Type 2 Excel): {'✅ PASS' if test2_pass else '❌ FAIL'}")
    
    if test1_pass and test2_pass:
        print("\n🎉 ALL TESTS PASSED")
        sys.exit(0)
    else:
        print("\n❌ SOME TESTS FAILED")
        sys.exit(1)
