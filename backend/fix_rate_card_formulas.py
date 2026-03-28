"""
Fix Rate Card SUM formulas after T&J column removal.
Move Hump from column G to column F, add SUM formula in column G.
"""

from openpyxl import load_workbook
from pathlib import Path

def fix_rate_card_formulas():
    """Fix Rate Card structure and formulas."""
    
    template_path = Path(__file__).parent / "reference_files" / "WestPark_FireDoor_CostSheet_v3_AlphaSights.xlsx"
    
    if not template_path.exists():
        print(f"❌ Template not found: {template_path}")
        return
    
    print(f"📂 Loading template: {template_path}")
    wb = load_workbook(template_path)
    
    if "Rate Card" not in wb.sheetnames:
        print("❌ Rate Card sheet not found")
        return
    
    rate_card = wb["Rate Card"]
    print("✅ Found Rate Card sheet")
    
    fixed_count = 0
    
    # Process all rows with codes (starting from row 6)
    for row in range(6, 50):
        code = rate_card.cell(row=row, column=1).value
        
        if code:  # Only process rows with a code
            # Current state:
            # D = Materials, E = Labour, F = 0 (T&J), G = Hump
            
            # Step 1: Move Hump from column G to column F
            hump_value = rate_card.cell(row=row, column=7).value  # Column G (old Hump)
            rate_card.cell(row=row, column=6).value = hump_value  # Column F (new Hump)
            
            # Step 2: Add SUM formula in column G (new Total)
            # Formula: =SUM(D{row}:F{row})
            formula = f"=SUM(D{row}:F{row})"
            rate_card.cell(row=row, column=7).value = formula
            
            # Verify calculation
            mat = rate_card.cell(row=row, column=4).value or 0
            lab = rate_card.cell(row=row, column=5).value or 0
            hump = hump_value or 0
            expected_total = mat + lab + hump
            
            print(f"  ✅ {code}: D={mat}, E={lab}, F={hump}, G={formula} → Total={expected_total}")
            fixed_count += 1
    
    # Save workbook
    wb.save(template_path)
    print(f"\n✅ Fixed {fixed_count} Rate Card rows")
    print(f"📁 Saved to: {template_path}")
    
    # Verify key rows
    print("\n" + "="*80)
    print("VERIFICATION")
    print("="*80)
    
    # Reload to verify
    wb = load_workbook(template_path, data_only=True)
    rate_card = wb["Rate Card"]
    
    verify_codes = ['A01', 'B01', 'B03']
    for code in verify_codes:
        for row in range(6, 50):
            if rate_card.cell(row=row, column=1).value == code:
                mat = rate_card.cell(row=row, column=4).value or 0
                lab = rate_card.cell(row=row, column=5).value or 0
                hump = rate_card.cell(row=row, column=6).value or 0
                total = rate_card.cell(row=row, column=7).value or 0
                print(f"{code}: {mat} + {lab} + {hump} = £{total:.2f}")
                break

if __name__ == "__main__":
    fix_rate_card_formulas()
