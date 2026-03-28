"""
Update Rate Card with Alex's real rates.
This replaces all mock rates with confirmed values from Alex.
"""

from openpyxl import load_workbook
from pathlib import Path

# Real rates from Alex
REAL_RATES = {
    # A-SERIES (Replacement Doors)
    'A01': {'materials': 593.08, 'labour': 250, 'humping': 40, 'note': ''},
    'A02': {'materials': 682.18, 'labour': 250, 'humping': 40, 'note': ''},
    'A03': {'materials': 770.18, 'labour': 250, 'humping': 40, 'note': ''},
    'A04': {'materials': 993.73, 'labour': 350, 'humping': 80, 'note': ''},
    'A05': {'materials': 593.08, 'labour': 250, 'humping': 40, 'note': ''},
    'A06': {'materials': 682.18, 'labour': 250, 'humping': 40, 'note': ''},
    'A07': {'materials': 770.18, 'labour': 250, 'humping': 40, 'note': ''},
    'A08': {'materials': 993.73, 'labour': 350, 'humping': 80, 'note': ''},
    'A09': {'materials': 768.65, 'labour': 250, 'humping': 40, 'note': ''},
    'A10': {'materials': 865.15, 'labour': 250, 'humping': 40, 'note': ''},
    'A11': {'materials': 1229.20, 'labour': 350, 'humping': 80, 'note': ''},
    'A12': {'materials': 100.00, 'labour': 0, 'humping': 0, 'note': ''},  # E/O Hardwood Veneer
    'A14': {'materials': 139.18, 'labour': 0, 'humping': 0, 'note': ''},  # E/O FD30 Vision Panel ≤300x600
    'A16': {'materials': 206.00, 'labour': 0, 'humping': 0, 'note': ''},  # E/O FD60 Vision Panel ≤300x600
    
    # B-SERIES (Remedial Works)
    'B01': {'materials': 67.65, 'labour': 75, 'humping': 0, 'note': ''},
    'B02': {'materials': 67.65, 'labour': 75, 'humping': 0, 'note': ''},
    'B03': {'materials': 45.00, 'labour': 37.50, 'humping': 0, 'note': ''},
    'B04': {'materials': 15.00, 'labour': 75, 'humping': 0, 'note': ''},
    'B05': {'materials': 25.00, 'labour': 75, 'humping': 0, 'note': ''},
    'B06': {'materials': 10.00, 'labour': 75, 'humping': 0, 'note': ''},
    'B07': {'materials': 10.00, 'labour': 37.50, 'humping': 0, 'note': ''},
    'B11': {'materials': 30.00, 'labour': 37.50, 'humping': 0, 'note': ''},
    
    # C-SERIES (Preliminaries)
    'C04a': {'materials': 0, 'labour': 150.00, 'humping': 0, 'note': ''},  # Out-of-hours weekday
    'C04b': {'materials': 0, 'labour': 300.00, 'humping': 0, 'note': ''},  # Out-of-hours Sunday
}

# Mock rates (pending confirmation from Alex/Matt)
MOCK_RATES = {
    'A13': {'materials': 150.00, 'labour': 0, 'humping': 0, 'note': 'Alex to confirm — under investigation'},
    'A15': {'materials': 200.00, 'labour': 0, 'humping': 0, 'note': 'Alex to confirm price for larger FD30 panel'},
    'A17': {'materials': 280.00, 'labour': 0, 'humping': 0, 'note': 'Alex to confirm price for larger FD60 panel'},
    'B08': {'materials': 325.00, 'labour': 0, 'humping': 0, 'note': 'Alex cannot price access control items — quote separately per job'},
    'B09': {'materials': 125.00, 'labour': 0, 'humping': 0, 'note': 'Alex cannot price access control items — quote separately per job'},
    'B10': {'materials': 300.00, 'labour': 0, 'humping': 0, 'note': 'Hourly rate £37.50/hr or minimum 1 day carpenter £300 cost — Matt to confirm per-door rate'},
    'B12': {'materials': 500.00, 'labour': 0, 'humping': 0, 'note': 'Hourly rate £62.50/hr or minimum 1 day magic man £500 cost — Matt to confirm per-door rate'},
    'C01': {'materials': 170.00, 'labour': 0, 'humping': 0, 'note': 'Alex to confirm call-out rate'},
    'C02': {'materials': 300.00, 'labour': 0, 'humping': 0, 'note': 'Alex: £300 minimum per visit — not per door. Matt to confirm how to allocate across jobs'},
    'C03': {'materials': 40.00, 'labour': 0, 'humping': 0, 'note': 'Alex to confirm per-door certification rate'},
    'C05': {'materials': 500.00, 'labour': 0, 'humping': 0, 'note': 'Job by job basis — Alex cannot provide standard rate'},
}

def update_rate_card():
    """Update Rate Card sheet with real and mock rates."""
    
    template_path = Path(__file__).parent / "reference_files" / "WestPark_FireDoor_CostSheet_v3_AlphaSights.xlsx"
    
    if not template_path.exists():
        print(f"❌ Template not found: {template_path}")
        return
    
    print(f"📂 Loading template: {template_path}")
    wb = load_workbook(template_path)
    
    if "Rate Card" not in wb.sheetnames:
        print("❌ Rate Card sheet not found in template")
        return
    
    rate_card = wb["Rate Card"]
    print("✅ Found Rate Card sheet")
    
    # Combine real and mock rates
    all_rates = {**REAL_RATES, **MOCK_RATES}
    
    updated_count = 0
    mock_count = 0
    
    # FIRST PASS: Clear old mock flags from column H (moved to column J)
    print("Clearing old mock flags from Column H...")
    h_cleared = 0
    for row in range(6, 50):
        code = rate_card.cell(row=row, column=1).value
        if code:  # Only process rows with a code
            # Clear column H if it contains text (old mock flags)
            col_h = rate_card.cell(row=row, column=8).value
            if col_h and isinstance(col_h, str) and 'MOCK' in col_h:
                rate_card.cell(row=row, column=8).value = None
                h_cleared += 1
    print(f"✅ Cleared {h_cleared} old mock flags from column H\n")
    
    # SECOND PASS: Update rates for known codes
    # Scan Rate Card sheet for codes (starting from row 6)
    for row in range(6, 50):  # Check up to row 50
        code = rate_card.cell(row=row, column=1).value  # Column A = Code
        
        if code and str(code).strip() in all_rates:
            code_str = str(code).strip()
            rates = all_rates[code_str]
            
            # Update Materials (Column D)
            rate_card.cell(row=row, column=4).value = rates['materials']
            
            # Update Labour (Column E)
            rate_card.cell(row=row, column=5).value = rates['labour']
            
            # Update Humping (Column F)
            rate_card.cell(row=row, column=6).value = rates['humping']
            
            # Add note if this is a mock rate (Column J = 10, avoid Column H which gets overwritten by processor)
            if code_str in MOCK_RATES and rates['note']:
                rate_card.cell(row=row, column=10).value = f"⚠️ MOCK: {rates['note']}"
                mock_count += 1
                print(f"  🟡 {code_str}: Materials=£{rates['materials']}, Labour=£{rates['labour']}, Humping=£{rates['humping']} (MOCK - {rates['note'][:50]}...)")
            else:
                # Clear any existing mock note for confirmed rates
                rate_card.cell(row=row, column=10).value = ""
                print(f"  ✅ {code_str}: Materials=£{rates['materials']}, Labour=£{rates['labour']}, Humping=£{rates['humping']}")
            
            updated_count += 1
    
    # THIRD PASS: Calculate and write Total column (Column G) for ALL rows
    print("\nCalculating totals for all Rate Card rows...")
    totals_written = 0
    for row in range(6, 50):
        code = rate_card.cell(row=row, column=1).value
        if code:  # Only process rows with a code
            mat = rate_card.cell(row=row, column=4).value or 0
            lab = rate_card.cell(row=row, column=5).value or 0
            hump = rate_card.cell(row=row, column=6).value or 0
            total = mat + lab + hump
            
            # Write numeric total to Column G
            rate_card.cell(row=row, column=7).value = total
            totals_written += 1
            
            if code in ['A01', 'A04', 'A09', 'B01', 'B03', 'B11']:  # Verification samples
                print(f"  ✅ {code}: {mat} + {lab} + {hump} = £{total:.2f}")
    
    print(f"✅ Wrote {totals_written} total values to Column G\n")
    
    # Save workbook
    wb.save(template_path)
    print(f"\n✅ Updated {updated_count} rates in Rate Card sheet")
    print(f"   - {updated_count - mock_count} confirmed rates")
    print(f"   - {mock_count} mock rates (pending confirmation)")
    print(f"📁 Saved to: {template_path}")
    
    # Generate summary for documentation
    print("\n" + "="*80)
    print("RATE CARD UPDATE SUMMARY")
    print("="*80)
    print("\nCONFIRMED RATES:")
    for code in sorted(REAL_RATES.keys()):
        rates = REAL_RATES[code]
        total = rates['materials'] + rates['labour'] + rates['humping']
        print(f"  {code}: Materials=£{rates['materials']}, Labour=£{rates['labour']}, Humping=£{rates['humping']} → Total=£{total:.2f}")
    
    print("\nMOCK RATES (PENDING CONFIRMATION):")
    for code in sorted(MOCK_RATES.keys()):
        rates = MOCK_RATES[code]
        total = rates['materials'] + rates['labour'] + rates['humping']
        print(f"  {code}: £{total:.2f} — {rates['note']}")

if __name__ == "__main__":
    update_rate_card()
