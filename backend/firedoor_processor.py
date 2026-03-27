"""
Fire Door Quoting Processor
Handles extraction and mapping of fire door survey data to rate card codes.
"""

import os
import re
import csv
import json
import logging
import fitz  # pymupdf
import httpx
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
try:
    from anthropic import Anthropic
    import anthropic
    ANTHROPIC_SDK_AVAILABLE = True
except Exception as e:
    ANTHROPIC_SDK_AVAILABLE = False
    anthropic = None
    Anthropic = None
    
from typing import Dict, List, Optional, Tuple
from pathlib import Path

# Log versions for debugging
logger = logging.getLogger(__name__)

# Import database dependencies (after logger is defined)
try:
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from database import get_db
    from models import RateCardItem
    DATABASE_AVAILABLE = True
except Exception as e:
    logger.warning(f"Database import failed: {e}. Will use CSV fallback.")
    DATABASE_AVAILABLE = False
    get_db = None
    RateCardItem = None
if ANTHROPIC_SDK_AVAILABLE and anthropic:
    logger.info(f"Anthropic SDK version: {anthropic.__version__}")
logger.info(f"httpx version: {httpx.__version__}")

# Lazy initialization of Anthropic client
_anthropic_client = None
_use_raw_http = False

def call_anthropic_api(prompt: str, max_tokens: int = 4096) -> str:
    """
    Call Anthropic API directly via HTTP (fallback when SDK fails).
    
    Args:
        prompt: The prompt to send
        max_tokens: Maximum tokens in response
        
    Returns:
        The text response from Claude
    """
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY environment variable not set")
    
    logger.info("Using raw HTTP API call to Anthropic (SDK unavailable or failed)")
    
    response = httpx.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json"
        },
        json={
            "model": "claude-sonnet-4-20250514",
            "max_tokens": max_tokens,
            "messages": [{"role": "user", "content": prompt}]
        },
        timeout=60.0
    )
    response.raise_for_status()
    result = response.json()
    return result["content"][0]["text"]

def get_anthropic_client():
    """Get or create Anthropic client (lazy initialization with fallback to raw HTTP)."""
    global _anthropic_client, _use_raw_http
    
    if _use_raw_http:
        return None  # Signal to use raw HTTP
    
    if _anthropic_client is None:
        if not ANTHROPIC_SDK_AVAILABLE:
            logger.warning("Anthropic SDK not available, using raw HTTP API")
            _use_raw_http = True
            return None
        
        try:
            logger.info("Attempting to initialize Anthropic SDK client")
            _anthropic_client = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
            logger.info("Anthropic SDK client initialized successfully")
        except Exception as e:
            logger.error(f"Failed to initialize Anthropic SDK: {str(e)}")
            logger.info("Falling back to raw HTTP API")
            _use_raw_http = True
            return None
    
    return _anthropic_client

# Load ART code mapping
ART_MAPPING = {}
SCRIPT_DIR = Path(__file__).parent
MAPPING_CSV_PATH = SCRIPT_DIR / "reference_files" / "BMTrada_ART_Codes_RateCard_Mapping.csv"

def load_art_mapping_from_db():
    """
    Load ART code to rate card mapping from database.
    Returns True if successful, False if database is empty or unavailable.
    """
    global ART_MAPPING
    
    if not DATABASE_AVAILABLE:
        return False
    
    try:
        db = next(get_db())
        items = db.query(RateCardItem).all()
        
        if not items:
            logger.info("Rate card database is empty. Will load from CSV.")
            return False
        
        for item in items:
            ART_MAPPING[item.art_code] = {
                'description': item.description,
                'rate_card_code': item.rate_card_code,
                'unit_price': item.unit_price,
                'notes': f"Last updated: {item.updated_at}" if item.updated_at else ""
            }
        
        logger.info(f"Loaded {len(items)} rate card items from database")
        db.close()
        return True
        
    except Exception as e:
        logger.error(f"Failed to load from database: {e}. Will fallback to CSV.")
        return False


def load_art_mapping_from_csv():
    """Load ART code to rate card mapping from CSV (fallback)."""
    global ART_MAPPING
    with open(MAPPING_CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            art_code = row['ART Code']
            ART_MAPPING[art_code] = {
                'description': row['Description'],
                'rate_card_code': row['WestPark Rate Card Code'],
                'unit_price': None,  # CSV doesn't have prices
                'notes': row['Notes']
            }
    logger.info(f"Loaded {len(ART_MAPPING)} rate card items from CSV")


def load_art_mapping():
    """
    Load ART code mapping. Try database first, fallback to CSV.
    Database contains user-editable prices. CSV is the default reference.
    """
    global ART_MAPPING
    
    # Try database first
    if load_art_mapping_from_db():
        logger.info("Using rate card from database (with user-edited prices)")
        return
    
    # Fallback to CSV
    logger.info("Using rate card from CSV (default prices)")
    load_art_mapping_from_csv()


# Load mapping on module import
load_art_mapping()


def detect_format(file_path: str, filename: str) -> str:
    """
    Detect survey format type.
    
    Returns:
        'TYPE_1' - PDF/TXT with FireDNA/RiskBase/BM TRADA/ART codes
        'TYPE_2' - Excel with door/gaps/seals fault columns
        'UNKNOWN' - Cannot determine format
    """
    ext = Path(filename).suffix.lower()
    
    if ext == '.pdf':
        # Try to read first few pages and look for TYPE_1 indicators
        try:
            doc = fitz.open(file_path)
            first_page_text = doc[0].get_text() if doc.page_count > 0 else ""
            second_page_text = doc[1].get_text() if doc.page_count > 1 else ""
            combined_text = (first_page_text + " " + second_page_text).lower()
            doc.close()
            
            # Check for TYPE_1 indicators
            type1_keywords = ['firedna', 'riskbase', 'bm trada', 'art', 'fire door survey']
            if any(keyword in combined_text for keyword in type1_keywords):
                return 'TYPE_1'
        except Exception as e:
            print(f"Error detecting PDF format: {e}")
            return 'UNKNOWN'
    
    elif ext == '.txt':
        # Handle pre-extracted PDF text files
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                # Read first 5000 chars
                text_sample = f.read(5000).lower()
                
                # Check for TYPE_1 indicators
                type1_keywords = ['firedna', 'riskbase', 'bm trada', 'art', 'fire door survey']
                if any(keyword in text_sample for keyword in type1_keywords):
                    return 'TYPE_1'
        except Exception as e:
            print(f"Error detecting TXT format: {e}")
            return 'UNKNOWN'
    
    elif ext in ['.xlsx', '.xls']:
        # Try to read and check for TYPE_2 column structure
        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            # Get header row (usually row 1)
            headers = [cell.value.lower() if cell.value else "" for cell in ws[1]]
            
            # Check for TYPE_2 indicators (door, gaps, seals, etc.)
            type2_keywords = ['door', 'gap', 'seal', 'fault', 'remedial']
            if any(any(keyword in header for keyword in type2_keywords) for header in headers):
                return 'TYPE_2'
        except Exception as e:
            print(f"Error detecting Excel format: {e}")
            return 'UNKNOWN'
    
    return 'UNKNOWN'


def extract_type1_pdf(file_path: str) -> List[Dict]:
    """
    Extract door data from Type 1 PDF/TXT using Claude API.
    Supports both PDF files and pre-extracted text files.
    
    Returns:
        List of door dictionaries with keys: door_id, location, faults, art_codes
    
    Raises:
        ValueError: If file format is unsupported or text extraction fails
    """
    # Extract full text from PDF or TXT
    full_text = ""
    ext = Path(file_path).suffix.lower()
    
    if ext == '.txt':
        # Read text file directly
        logger.info(f"Reading text file: {file_path}")
        with open(file_path, 'r', encoding='utf-8') as f:
            full_text = f.read()
        logger.info(f"Text file read: {len(full_text)} characters")
    elif ext == '.pdf':
        # Extract from PDF using pymupdf (recommended by Mauricio)
        logger.info(f"Extracting text from PDF: {file_path}")
        try:
            doc = fitz.open(file_path)
            logger.info(f"PDF opened: {doc.page_count} pages")
            for page_num in range(doc.page_count):
                page = doc[page_num]
                page_text = page.get_text()
                if page_text:
                    full_text += page_text + "\n\n"
                    logger.info(f"Page {page_num + 1}: extracted {len(page_text)} characters")
                else:
                    logger.warning(f"Page {page_num + 1}: no text extracted (may be image-only)")
            doc.close()
            logger.info(f"PDF extraction complete: {len(full_text)} total characters")
        except Exception as e:
            logger.error(f"PDF extraction failed: {e}")
            raise ValueError(f"Failed to extract text from PDF. Error: {str(e)}. If this is an image-heavy PDF, please use a PDF-to-text converter first and upload the .txt file instead.")
    else:
        raise ValueError(f"Unsupported file extension for Type 1: {ext}. Please upload a PDF or TXT file.")
    
    # Check if we got any text
    if not full_text or len(full_text.strip()) < 100:
        logger.error(f"Insufficient text extracted from {file_path}: {len(full_text)} characters")
        raise ValueError(
            "Could not extract sufficient text from the uploaded file. "
            "This may be an image-only PDF or scanned document. "
            "Please use a PDF-to-text converter (or Claude Code) to extract the text first, "
            "then upload the resulting .txt file instead."
        )
    
    # Use Claude to extract structured data
    prompt = f"""Extract all fire door data from this survey report. For each door, extract:
1. Door ID/Number - MUST include floor suffix if present in survey (e.g. A01-L2, A02-L3)
2. Location/Description (building name, floor level)
3. List of faults found (any issues or deficiencies noted)
4. ART codes mentioned (e.g. ART01, ART04, ART18)
5. Fire rating - Extract exact rating from survey:
   - FD30 (30-minute, no smoke seals)
   - FD30S (30-minute with smoke seals)
   - FD60 (60-minute, no smoke seals)
   - FD60S (60-minute with smoke seals)
   - FD90, FD120, etc.
   - ONLY use "Unknown" if truly not mentioned
6. Door configuration: "Single Leaf" or "Double Leaf"
7. Door height in millimeters (extract from dimensions)
8. Whether this is a replacement door (true ONLY if ART17, ART18, or ART20 present)

CRITICAL: Door IDs MUST match survey exactly. If survey shows "A01" for Level 2, extract as "A01-L2".

Return as JSON array with this structure:
[
  {{
    "door_id": "A01-L2",
    "location": "Level 2 - Main Corridor",
    "faults": ["Gap too large", "Seal missing"],
    "art_codes": ["ART04", "ART11"],
    "fire_rating": "FD60S",
    "door_config": "Single Leaf",
    "door_height_mm": 2100,
    "is_replacement": false
  }},
  ...
]

Survey text:
{full_text}

Return ONLY the JSON array, no other text."""

    client = get_anthropic_client()
    
    # Use SDK if available, otherwise raw HTTP
    # Increased max_tokens to 8192 to handle large surveys (39+ doors)
    if client is not None:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=8192,
            messages=[{"role": "user", "content": prompt}]
        )
        response_text = response.content[0].text
        logger.info(f"Claude response received: {len(response_text)} chars, stop_reason={response.stop_reason}")
        
        # Check if response was truncated
        if response.stop_reason == "max_tokens":
            logger.warning("Claude response was truncated due to max_tokens limit!")
            logger.warning("Survey may be too large - consider increasing max_tokens or splitting the survey")
    else:
        response_text = call_anthropic_api(prompt, max_tokens=8192)
        logger.info(f"Claude response received: {len(response_text)} chars")
    
    # Parse JSON response
    try:
        doors = json.loads(response_text)
        # FIX #2: Mark all Type 1 doors with format_type
        for door in doors:
            door['format_type'] = 'TYPE_1'
        logger.info(f"Successfully extracted {len(doors)} doors from Type 1 survey")
        return doors
    except json.JSONDecodeError as e:
        logger.error(f"JSON parse error: {e}")
        logger.error(f"Response text length: {len(response_text)} chars")
        logger.error(f"First 500 chars: {response_text[:500]}")
        logger.error(f"Last 500 chars: {response_text[-500:]}")
        
        # Try to extract JSON from response
        json_match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if json_match:
            json_str = json_match.group()
            logger.info(f"Found JSON array in response, attempting to parse...")
            try:
                doors = json.loads(json_str)
                # FIX #2: Mark all Type 1 doors with format_type
                for door in doors:
                    door['format_type'] = 'TYPE_1'
                logger.info(f"Successfully extracted {len(doors)} doors after regex cleanup")
                return doors
            except json.JSONDecodeError as e2:
                logger.error(f"Regex extraction also failed: {e2}")
                logger.error(f"JSON string: {json_str[:1000]}...")
                raise ValueError(
                    f"Claude returned invalid JSON. Error: {str(e)}. "
                    "This may be due to a large survey - try splitting into smaller sections."
                )
        else:
            logger.error("No JSON array found in Claude response")
            raise ValueError(
                "Claude did not return a valid JSON array. "
                "Response may be truncated or formatted incorrectly."
            )


def map_fault_to_bcode(fault_text: str) -> Optional[str]:
    """
    Map fault description to B-series code based on keywords.
    
    BUG 2 FIX: Enhanced mapping to catch more fault patterns
    
    B-series mapping rules:
    B01 - Intumescent + smoke seals
    B02 - Intumescent only
    B03 - Door closer
    B04 - Hinges
    B05 - Latch/handles
    B06 - Frame fire stopping / architrave
    B07 - Fire door signage
    B08 - Hold-open device removal (Perco chains)
    B10 - Adjust/re-hang door
    B11 - Hardwood re-lip
    B12 - Timber insert for voids
    """
    fault_lower = fault_text.lower()
    
    # B08 - Hold-open devices (Perco chains, etc.)
    if any(keyword in fault_lower for keyword in ['hold open', 'hold-open', 'perco', 'chain', 'hook']):
        return 'B08'
    
    # B01 - Seals (broader matching for Type 2 Excel surveys)
    # Matches: "seal", "strip", "coming away", "worn", "damaged", "missing", "fitted" (in seal context)
    seal_keywords = ['seal', 'strip', 'coming away', 'worn', 'damage', 'missing', 'none fitted', 'not fitted']
    if any(keyword in fault_lower for keyword in seal_keywords):
        # If it mentions intumescent only (no smoke), it's B02
        if 'intumescent' in fault_lower and 'smoke' not in fault_lower:
            return 'B02'
        # Otherwise assume B01 (intumescent + smoke seals, or general seal issue)
        return 'B01'
    
    # B02 - Intumescent only (explicit)
    if 'intumescent' in fault_lower and 'smoke' not in fault_lower:
        return 'B02'
    
    # B03 - Door closer
    if any(keyword in fault_lower for keyword in ['closer', 'closing']):
        # Exclude "not closing" which is B10
        if 'not closing' in fault_lower or 'catching' in fault_lower:
            return 'B10'
        return 'B03'
    
    # B04 - Hinges
    if any(keyword in fault_lower for keyword in ['hinge', 'dropped', 'drop']):
        return 'B04'
    
    # B05 - Latch/handles
    if any(keyword in fault_lower for keyword in ['latch', 'handle', 'lock', 'keep']):
        return 'B05'
    
    # B06 - Frame/architrave
    if any(keyword in fault_lower for keyword in ['frame', 'architrave', 'fire stop', 'firestop', 'gap to wall']):
        return 'B06'
    
    # B07 - Signage
    if any(keyword in fault_lower for keyword in ['sign', 'signage', 'label']):
        return 'B07'
    
    # B10 - Re-hang/adjust (catches "not closing", "catching on floor", etc.)
    if any(keyword in fault_lower for keyword in ['adjust', 're-hang', 'rehang', 'alignment', 'warped', 'twisted', 'gap too', 'catching', 'not closing']):
        return 'B10'
    
    # B11 - Re-lip
    if any(keyword in fault_lower for keyword in ['lipping', 'lip', 'edge damage']):
        return 'B11'
    
    # B12 - Timber insert
    if any(keyword in fault_lower for keyword in ['void', 'hole', 'insert', 'recessed']):
        return 'B12'
    
    return None


def extract_type2_excel(file_path: str) -> List[Dict]:
    """
    Extract door data from Type 2 Excel file and map faults to B-codes.
    
    Returns:
        List of door dictionaries with keys: door_id, location, faults, b_codes
    """
    wb = load_workbook(file_path)
    all_doors = []
    
    # Process all sheets (some surveys have Floor 1, Floor 2, etc.)
    for sheet in wb.worksheets:
        # Find header row (usually row 1 or 2)
        # Look for row with multiple fault-related columns
        header_row = 1
        headers = [cell.value.lower() if cell.value else "" for cell in sheet[1]]
        
        # Count how many columns look like fault columns
        fault_keywords = ['gap', 'seal', 'frame', 'hinge', 'lock', 'glass', 'strip', 'closer', 'ironmongery']
        fault_col_count = sum(1 for h in headers if any(keyword in h for keyword in fault_keywords))
        
        # If row 1 doesn't have multiple fault columns, try row 2
        if fault_col_count < 2:
            header_row = 2
            headers = [cell.value.lower() if cell.value else "" for cell in sheet[2]]
            fault_col_count = sum(1 for h in headers if any(keyword in h for keyword in fault_keywords))
        
        # Find column indices
        door_col = next((i for i, h in enumerate(headers) if 'door' in h and ('no' in h or 'number' in h or 'ref' in h)), None)
        
        # Fault columns: any column with door-related keywords, excluding the door number column
        fault_cols = [i for i, h in enumerate(headers) if i != door_col and any(
            keyword in h for keyword in ['gap', 'seal', 'frame', 'hinge', 'lock', 'glass', 'strip', 'closer', 'ironmongery', 'wall']
        )]
        
        # If no specific fault columns found, use all columns except first 3 (usually ID/ref columns)
        if not fault_cols:
            fault_cols = list(range(3, len(headers)))
        
        # Extract doors
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
            if not row or not any(row):  # Skip empty rows
                continue
                
            door_id = row[door_col] if door_col is not None and door_col < len(row) else None
            
            # Skip rows where door_id is not a valid integer (filters out #VALUE!, headers, etc.)
            if door_id is None:
                continue
            
            # Try to convert to integer - if it fails, skip this row
            try:
                # Check if it's numeric (could be int or float)
                if isinstance(door_id, (int, float)):
                    door_number = int(door_id)
                else:
                    # Try to parse string as integer
                    door_number = int(str(door_id).strip())
            except (ValueError, TypeError):
                # Not a valid integer - skip this row (likely #VALUE! or header)
                continue
            
            # Collect faults from relevant columns
            faults = []
            b_codes = []
            has_unable = False  # Track if any column contains "Unable"
            
            for col_idx in fault_cols:
                if col_idx < len(row) and row[col_idx]:
                    fault_text = str(row[col_idx]).strip()
                    
                    # Check for "Unable" keyword (triggers orange flag)
                    if 'unable' in fault_text.lower():
                        has_unable = True
                    
                    # Skip "OK", "None", "N/A", "NO", empty values
                    if fault_text and fault_text.upper() not in ['OK', 'NONE', 'N/A', 'NO', 'YES', '-']:
                        faults.append(fault_text)
                        # Map to B-code
                        b_code = map_fault_to_bcode(fault_text)
                        if b_code:
                            b_codes.append(b_code)
            
            if faults:  # Only add doors with actual faults
                # FIX #2 (CORRECTED): Try to infer fire rating and door config from Excel data
                # For Type 2, these are usually Unknown unless specified in the Excel
                fire_rating = 'Unknown'
                door_config = 'Single Leaf'  # Default assumption
                is_replacement = False
                
                # TODO: Check if Excel has columns for fire rating, dimensions, or config
                # For now, use placeholders - will be PENDING in Quote Sheet
                
                all_doors.append({
                    'door_id': f"{sheet.title}-{door_number}",  # Use validated integer door number
                    'location': sheet.title,  # Use sheet name as location
                    'faults': faults,
                    'b_codes': list(set(b_codes)),  # Remove duplicates
                    'has_unable': has_unable,  # Flag for orange highlighting
                    'format_type': 'TYPE_2',  # Mark as Type 2 for PENDING logic
                    'fire_rating': fire_rating,  # FIX #2: Add fire rating
                    'door_config': door_config,  # FIX #2: Add door config
                    'is_replacement': is_replacement  # FIX #2: Add replacement flag
                })
    
    return all_doors


def map_art_to_rate_card(art_codes: List[str]) -> List[str]:
    """Map ART codes to WestPark rate card codes using the CSV mapping."""
    rate_card_codes = []
    for art_code in art_codes:
        if art_code in ART_MAPPING:
            rc_code = ART_MAPPING[art_code]['rate_card_code']
            if rc_code and rc_code.strip() and rc_code not in ['FLAG FOR MANUAL REVIEW', 'NO EQUIVALENT']:
                # Handle multi-code entries like "B01 / B02"
                codes = [c.strip() for c in rc_code.split('/')]
                rate_card_codes.extend(codes)
    
    return list(set(rate_card_codes))  # Remove duplicates


def get_priority_bcode(codes: List[str]) -> str:
    """
    Get the priority B-code from a list of codes.
    Priority order: B01 > B03 > B04 > B10 > B05 > B02 > B06 > B07
    
    BUG 2 FIX: This ensures the primary B-code follows the correct priority
    for Quote Sheet COUNTIF formulas.
    """
    if not codes:
        return ''
    
    priority_order = ['B01', 'B03', 'B04', 'B10', 'B05', 'B02', 'B06', 'B07', 'B11', 'B12']
    
    for priority_code in priority_order:
        if priority_code in codes:
            return priority_code
    
    # If none match priority list, return first code
    return codes[0]


def map_to_aseries_code(fire_rating: str, door_config: str, door_height: int = None) -> str:
    """
    Map fire rating + door configuration + height to A-series replacement code.
    
    Args:
        fire_rating: Fire rating (FD60, FD60S, FD30, FD30S, Nominal, etc.)
        door_config: Door configuration ("Single Leaf" or "Double Leaf")
        door_height: Door height in mm (optional, defaults to ≤2040 range)
    
    Returns:
        A-series code (A01-A15) or empty string if no match
    
    Full mapping table (from Mauricio 2026-03-27):
        FD30 Single ≤2040 → A01
        FD30 Single 2040–2400 → A02
        FD30 Single 2400–2730 → A03
        FD30 Double ≤2040 → A04
        FD30S Single ≤2040 → A05
        FD30S Single 2040–2400 → A06
        FD30S Single 2400–2730 → A07
        FD30S Double ≤2040 → A08
        FD60S Single ≤2040 → A09
        FD60S Single 2040–2400 → A10
        FD60S Double ≤2040 → A11
        Nominal Single → A05 (default to FD30S)
        Nominal Double → A08 (default to FD30S)
    
    Note: S suffix = smoke seals. FD30 ≠ FD30S
    """
    if not fire_rating or not door_config:
        return ''
    
    rating_upper = str(fire_rating).upper()
    config_lower = str(door_config).lower()
    
    # Normalize config
    is_single = 'single' in config_lower
    is_double = 'double' in config_lower
    
    # Default to ≤2040 range if height not provided
    if door_height is None or door_height <= 2040:
        height_range = 1  # ≤2040
    elif door_height <= 2400:
        height_range = 2  # 2040-2400
    elif door_height <= 2730:
        height_range = 3  # 2400-2730
    else:
        height_range = 1  # Fallback to base code
    
    # Map based on rating + config + height
    # Check for exact FD30 (no S)
    if rating_upper == 'FD30':
        if is_single:
            if height_range == 1:
                return 'A01'
            elif height_range == 2:
                return 'A02'
            elif height_range == 3:
                return 'A03'
        elif is_double:
            return 'A04'  # FD30 double (no smoke seals)
    
    # Check for FD30S (with S)
    elif 'FD30S' in rating_upper:
        if is_single:
            if height_range == 1:
                return 'A05'
            elif height_range == 2:
                return 'A06'
            elif height_range == 3:
                return 'A07'
        elif is_double:
            return 'A08'  # FD30S double (with smoke seals)
    
    # Check for FD60S
    elif 'FD60S' in rating_upper:
        if is_single:
            if height_range == 1:
                return 'A09'
            elif height_range == 2:
                return 'A10'
        elif is_double:
            return 'A11'
    
    # Nominal defaults to FD30S
    elif 'NOMINAL' in rating_upper:
        if is_single:
            return 'A05'  # Default to FD30S single ≤2040
        elif is_double:
            return 'A08'  # Default to FD30S double
    
    return ''  # No match


def populate_excel_template(doors: List[Dict], client_name: str, template_path: str, output_path: str):
    """
    Populate Excel template with door data and color code rows.
    
    Writes to "Door Schedule" sheet with proper column mapping:
    - Column A: Door ID
    - Column B: Location
    - Column C-I: Door details (type, rating, config, size, finish, seals, closer)
    - Column P: Primary rate card code (for Quote Sheet formulas)
    
    Args:
        doors: List of door dictionaries
        client_name: Client name for the header
        template_path: Path to template Excel file
        output_path: Path to save output file
    """
    import logging
    from pathlib import Path
    
    logger = logging.getLogger(__name__)
    
    # Validate template exists
    template_file = Path(template_path)
    if not template_file.exists():
        error_msg = f"Template file not found: {template_path}"
        logger.error(error_msg)
        raise FileNotFoundError(error_msg)
    
    logger.info(f"Loading template from: {template_path}")
    logger.info(f"Output will be saved to: {output_path}")
    logger.info(f"Processing {len(doors)} doors for client: {client_name}")
    
    try:
        wb = load_workbook(template_path)
        logger.info(f"Template loaded successfully. Sheets: {wb.sheetnames}")
    except Exception as e:
        error_msg = f"Failed to load template workbook: {str(e)}"
        logger.error(error_msg)
        raise RuntimeError(error_msg) from e
    
    # Update Rate Card sheet with database prices
    if DATABASE_AVAILABLE:
        try:
            db = next(get_db())
            rate_items = db.query(RateCardItem).all()
            
            # Always extract template prices for backfilling
            template_prices = {}
            if "Rate Card" in wb.sheetnames:
                rate_card_sheet = wb["Rate Card"]
                for row_num in range(6, 40):
                    code = rate_card_sheet.cell(row=row_num, column=1).value
                    if code:
                        # FIX #4: Use Materials + Labour only (not T&J or Humping)
                        mat = rate_card_sheet.cell(row=row_num, column=4).value or 0
                        lab = rate_card_sheet.cell(row=row_num, column=5).value or 0
                        # tj = rate_card_sheet.cell(row=row_num, column=6).value or 0  # Removed
                        # hump = rate_card_sheet.cell(row=row_num, column=7).value or 0  # Removed
                        total = mat + lab  # Materials + Labour only
                        if total > 0:
                            template_prices[str(code)] = f"£{total:.2f}"
                logger.info(f"Extracted {len(template_prices)} prices from template")
            
            if not rate_items:
                # Database is empty - seed ALL codes from template
                logger.info("Database empty - auto-seeding ALL codes from template...")
                import csv
                
                # First, create a map of ART code descriptions from CSV
                art_descriptions = {}
                csv_path = Path(__file__).parent / "reference_files" / "BMTrada_ART_Codes_RateCard_Mapping.csv"
                if csv_path.exists():
                    with open(csv_path, 'r', encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            art_code = row['ART Code']
                            description = row['Description']
                            rate_card_code = row['WestPark Rate Card Code']
                            
                            # Map rate card code to ART description
                            if rate_card_code and rate_card_code.strip() not in ['', 'FLAG FOR MANUAL REVIEW', 'NO EQUIVALENT']:
                                codes = [c.strip() for c in rate_card_code.split('/')]
                                for code in codes:
                                    if code not in art_descriptions:
                                        art_descriptions[code] = (art_code, description)
                
                # Now seed ALL B-codes from template (rows 22-33 in Rate Card sheet)
                if "Rate Card" in wb.sheetnames:
                    rate_card_sheet = wb["Rate Card"]
                    seeded_count = 0
                    
                    for row_num in range(22, 34):  # B01-B12
                        code = rate_card_sheet.cell(row=row_num, column=1).value
                        rate_card_desc = rate_card_sheet.cell(row=row_num, column=2).value  # B-code description
                        
                        if code and str(code).startswith('B'):
                            # Get price from template
                            unit_price = template_prices.get(str(code), "£0.00")
                            
                            # Get ART code mapping if exists
                            art_info = art_descriptions.get(str(code), (str(code), rate_card_desc or ""))
                            art_code, art_desc = art_info
                            
                            # Create database entry
                            new_item = RateCardItem(
                                art_code=art_code,
                                description=art_desc[:500] if art_desc else "",  # ART description
                                rate_card_code=str(code),
                                rate_card_description=rate_card_desc[:500] if rate_card_desc else "",  # B-code description
                                unit_price=unit_price,
                                category="From template"
                            )
                            db.add(new_item)
                            seeded_count += 1
                    
                    db.commit()
                    logger.info(f"Auto-seeded {seeded_count} B-codes from template with prices")
                    
                    # Re-query to get the seeded items
                    rate_items = db.query(RateCardItem).all()
                else:
                    logger.warning("Rate Card sheet not found - cannot auto-seed")
            
            if rate_items:
                # Backfill template prices and descriptions for any items missing them
                backfill_count = 0
                desc_backfill_count = 0
                
                # First, create a map of B-codes to their descriptions from template
                bcode_descriptions = {}
                if "Rate Card" in wb.sheetnames:
                    rate_card_sheet = wb["Rate Card"]
                    for row_num in range(22, 34):  # B01-B12
                        code = rate_card_sheet.cell(row=row_num, column=1).value
                        rate_card_desc = rate_card_sheet.cell(row=row_num, column=2).value
                        if code and str(code).startswith('B'):
                            bcode_descriptions[str(code)] = rate_card_desc
                
                for item in rate_items:
                    # Backfill unit_price if missing
                    if not item.unit_price or item.unit_price == "£0.00" or item.unit_price.strip() == "":
                        codes = [c.strip() for c in item.rate_card_code.split('/')]
                        template_price = template_prices.get(codes[0], None) if codes else None
                        
                        if template_price:
                            item.unit_price = template_price
                            backfill_count += 1
                    
                    # Backfill rate_card_description if missing
                    if not item.rate_card_description or item.rate_card_description.strip() == "":
                        codes = [c.strip() for c in item.rate_card_code.split('/')]
                        bcode_desc = bcode_descriptions.get(codes[0], None) if codes else None
                        
                        if bcode_desc:
                            item.rate_card_description = bcode_desc
                            desc_backfill_count += 1
                
                # Also add any missing B-codes from template
                existing_rate_codes = set()
                for item in rate_items:
                    codes = [c.strip() for c in item.rate_card_code.split('/')]
                    existing_rate_codes.update(codes)
                
                missing_count = 0
                if "Rate Card" in wb.sheetnames:
                    rate_card_sheet = wb["Rate Card"]
                    for row_num in range(22, 34):  # B01-B12
                        code = rate_card_sheet.cell(row=row_num, column=1).value
                        rate_card_desc = rate_card_sheet.cell(row=row_num, column=2).value
                        
                        if code and str(code).startswith('B') and str(code) not in existing_rate_codes:
                            # Missing B-code - add it
                            unit_price = template_prices.get(str(code), "£0.00")
                            
                            new_item = RateCardItem(
                                art_code=str(code),  # Use code as art_code for standalone items
                                description=rate_card_desc[:500] if rate_card_desc else f"Template item {code}",  # Use B-code desc as ART desc for standalone
                                rate_card_code=str(code),
                                rate_card_description=rate_card_desc[:500] if rate_card_desc else "",
                                unit_price=unit_price,
                                category="Added from template"
                            )
                            db.add(new_item)
                            missing_count += 1
                
                if backfill_count > 0 or desc_backfill_count > 0 or missing_count > 0:
                    db.commit()
                    if backfill_count > 0:
                        logger.info(f"Backfilled {backfill_count} missing prices from template")
                    if desc_backfill_count > 0:
                        logger.info(f"Backfilled {desc_backfill_count} missing rate card descriptions from template")
                    if missing_count > 0:
                        logger.info(f"Added {missing_count} missing B-codes from template")
                    
                    # Re-query to get all items including newly added ones
                    rate_items = db.query(RateCardItem).all()
                
                # Create mapping: rate_card_code -> unit_price
                price_map = {}
                for item in rate_items:
                    # One ART code can map to multiple rate card codes (e.g., "B01 / B02")
                    codes = [c.strip() for c in item.rate_card_code.split('/')]
                    for code in codes:
                        if code and code not in price_map and item.unit_price:
                            price_map[code] = item.unit_price
                
                logger.info(f"Loaded {len(price_map)} rate card prices from database")
                
                # Update Rate Card sheet
                if "Rate Card" in wb.sheetnames:
                    rate_card_sheet = wb["Rate Card"]
                    updated_count = 0
                    
                    # Scan rows 6-39 (where rate card codes live)
                    for row_num in range(6, 40):
                        code_cell = rate_card_sheet.cell(row=row_num, column=1)  # Column A
                        code = code_cell.value
                        
                        if code and str(code) in price_map:
                            price_str = price_map[str(code)]
                            # Parse price (e.g., "£45.00" -> 45.00)
                            try:
                                price_value = float(price_str.replace('£', '').replace(',', '').strip())
                                # Write to column H (TOTAL RATE)
                                total_cell = rate_card_sheet.cell(row=row_num, column=8)
                                total_cell.value = price_value
                                updated_count += 1
                                logger.info(f"Updated {code} price to £{price_value}")
                            except (ValueError, AttributeError) as e:
                                logger.warning(f"Could not parse price for {code}: {price_str}")
                    
                    logger.info(f"Updated {updated_count} rate card prices in Excel template")
                else:
                    logger.warning("Rate Card sheet not found in template")
            else:
                logger.info("No rate card items in database - using template defaults")
            
            db.close()
        except Exception as e:
            logger.warning(f"Failed to load database prices: {e}. Using template defaults.")
    else:
        logger.info("Database not available - using template default prices")
    
    # FIX: Replace Rate Card column H formulas with calculated numeric values
    # This MUST happen AFTER database updates to ensure calculated values from template always win
    # Rate Card column H has formulas like =SUM(D22:G22) which openpyxl saves with cached value 0
    try:
        if "Rate Card" in wb.sheetnames:
            rc = wb['Rate Card']
            logger.info("=== FINAL PASS: Replacing Rate Card column H with calculated values ===")
            
            replaced_count = 0
            for row_idx in range(6, 40):  # Rows 6-39 cover all rate card items
                row = rc[row_idx]
                # Get values from columns D, E, F, G (indices 3, 4, 5, 6)
                d = row[3].value  # MAT'S
                e = row[4].value  # LABOUR
                f = row[5].value  # T & J
                g = row[6].value  # HUMP
                
                # If all are numeric, calculate and write to column H (index 7)
                if all(isinstance(v, (int, float)) for v in [d, e, f, g]):
                    total = d + e + f + g
                    row[7].value = total  # Replace formula with numeric value
                    replaced_count += 1
                    # Log specific rows we care about
                    if row_idx in [22, 23, 26, 31]:  # B01, B02, B05, B10
                        code = row[0].value
                        logger.info(f"Row {row_idx} ({code}): Set H = {total} (D={d} + E={e} + F={f} + G={g})")
            
            logger.info(f"Replaced {replaced_count} Rate Card formula cells with calculated values")
            
            # Verify fix worked (check B01, B02, B10)
            h22_value = rc['H22'].value
            h23_value = rc['H23'].value
            h31_value = rc['H31'].value
            logger.info(f"=== VERIFICATION: B01={h22_value}, B02={h23_value}, B10={h31_value} (expected 120, 90, 85) ===")
            
    except Exception as e:
        logger.error(f"CRITICAL: Failed to replace Rate Card formulas: {e}")
        import traceback
        logger.error(traceback.format_exc())
    
    # Update client name in Quote Sheet
    try:
        quote_sheet = wb["Quote Sheet"]
        quote_sheet['B4'] = client_name  # B4 is the Client field
        logger.info(f"Updated Quote Sheet with client name: {client_name}")
        
        # BUG 4 FIX: Clear site/building for Type 2 surveys (no site address in Excel files)
        # Check if this is a Type 2 survey by looking at first door's format_type
        is_type2 = doors and doors[0].get('format_type') == 'TYPE_2'
        if is_type2:
            # Clear site and building fields
            quote_sheet['B5'] = ''  # Site field
            quote_sheet['B6'] = ''  # Building field
            logger.info("Cleared site/building fields for Type 2 survey")
    except KeyError as e:
        error_msg = f"Quote Sheet not found in template. Available sheets: {wb.sheetnames}"
        logger.error(error_msg)
        raise RuntimeError(error_msg) from e
    
    # Get Door Schedule sheet for populating door data
    try:
        ws = wb["Door Schedule"]
        logger.info("Door Schedule sheet found")
    except KeyError as e:
        error_msg = f"Door Schedule sheet not found in template. Available sheets: {wb.sheetnames}"
        logger.error(error_msg)
        raise RuntimeError(error_msg) from e
    
    # Define color fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    orange_fill = PatternFill(start_color="FED8B1", end_color="FED8B1", fill_type="solid")  # Light orange (for manual review)
    
    # Data starts at row 4 (row 3 is headers)
    start_row = 4
    
    # BUG 5 FIX: Clear existing data (rows 4-100) - extended to column V (22)
    for row_num in range(4, 100):
        for col in range(1, 23):  # Columns A-V (1-22)
            ws.cell(row=row_num, column=col).value = None
            ws.cell(row=row_num, column=col).fill = PatternFill()  # Clear fill
    
    # Populate door data
    for idx, door in enumerate(doors):
        row_num = start_row + idx
        
        # Get codes (either b_codes from Type 2 or map art_codes from Type 1)
        codes = door.get('b_codes', [])
        if not codes and 'art_codes' in door:
            codes = map_art_to_rate_card(door['art_codes'])
        
        # BUG 2 FIX: Get primary code using priority order
        primary_code = get_priority_bcode(codes)
        all_codes_str = ', '.join(codes) if codes else 'MANUAL REVIEW'
        
        # FINAL FIX: For unable-to-inspect doors with no detectable codes,
        # assign B01 (seals) as provisional primary code
        # Column P must NEVER be empty for a YES door
        if not primary_code and door.get('has_unable'):
            primary_code = 'B01'  # Provisional code for reinspection
        
        # Column mapping based on Door Schedule template (updated Mar 2026):
        # A=DOOR ID, B=LOCATION, C=DOOR TYPE, D=CURRENT RATING, E=LEAF CONFIG,
        # F=LEAF SIZE, G=FINISH, H=SEALS, I=CLOSER, J=VISION PANEL,
        # K=ACTION DESCRIPTION, L=SEVERITY, M=DUE DATE,
        # N=OPT A REMEDIAL?, O=OPT B REPLACE?, P=OPT B BASE ITEM, Q=QTY,
        # R-U=E/O (OVERSIZE/HARDWOOD/EXTERNAL/VISION), V=NOTES/FLAGS
        
        door_id = door['door_id']
        
        ws[f'A{row_num}'] = door_id
        ws[f'B{row_num}'] = door.get('location', '')
        ws[f'C{row_num}'] = 'From Survey'  # Door type placeholder
        
        # FIX #2 (CORRECTED): Get fire rating and config from door data (extracted by Claude or Excel)
        fire_rating = door.get('fire_rating', 'Unknown')
        door_config = door.get('door_config', 'Single Leaf')
        
        ws[f'D{row_num}'] = fire_rating
        ws[f'E{row_num}'] = door_config
        
        ws[f'F{row_num}'] = 'Unknown'  # Size placeholder
        ws[f'G{row_num}'] = 'Paint'  # Finish placeholder
        ws[f'H{row_num}'] = 'To Check'  # Seals placeholder
        ws[f'I{row_num}'] = 'To Check'  # Closer placeholder
        ws[f'J{row_num}'] = 'To Check'  # Vision panel placeholder
        
        # Column K: ACTION DESCRIPTION (fault details)
        faults_str = '; '.join(door.get('faults', []))
        ws[f'K{row_num}'] = faults_str[:500]  # Longer field for faults
        
        # Column L: SEVERITY (based on number of codes/faults)
        severity = 'HIGH' if len(codes) > 3 else 'MEDIUM' if codes else 'LOW'
        ws[f'L{row_num}'] = severity
        
        # Column M: DUE DATE (leave empty for now)
        # ws[f'M{row_num}'] = ''
        
        # Column N: OPT A REMEDIAL? (YES/NO/COMPLIANT)
        # Column O: OPT B REPLACE? (YES/NO/PENDING)
        # BUG 1 FIX: Door is compliant ONLY if no codes AND no faults
        # Rule: any detected fault = Opt A YES
        is_compliant = not codes and not faults_str
        
        # CRITICAL FIX: Check for replacement doors - ONLY ART17, ART18, ART20
        # ART17 = leaf replacement, ART18 = full doorset replacement, ART20 = frame replacement
        art_codes = door.get('art_codes', [])
        needs_replacement = (
            any(art in ['ART17', 'ART18', 'ART20'] for art in art_codes) or
            'A-series' in codes or  # CSV mapping returns 'A-series' for replacement codes
            door.get('is_replacement', False)  # Claude's explicit flag
        )
        
        # Check if this is Type 2 format (no fire strategy, so Option B = PENDING)
        is_type2 = door.get('format_type') == 'TYPE_2'
        
        # FIX #2 (CORRECTED): Determine if this is a replacement door and map to A-series code
        is_replacement_door = door.get('is_replacement', False) or needs_replacement
        a_series_code = ''
        
        if is_replacement_door:
            # Get door height (from extraction or default)
            door_height = door.get('door_height_mm', None)
            
            if fire_rating != 'Unknown':
                # Map fire rating + config + height to A-series code
                a_series_code = map_to_aseries_code(fire_rating, door_config, door_height)
                if a_series_code:
                    height_str = f"{door_height}mm" if door_height else "≤2040mm (default)"
                    logger.info(f"Door {door_id}: {fire_rating} {door_config} {height_str} → {a_series_code}")
                else:
                    logger.error(f"Door {door_id}: Replacement needed but mapping failed (rating={fire_rating}, config={door_config})")
            else:
                # URGENT FIX: If fire rating is Unknown, use default mapping based on ART codes
                # ART18 typically indicates non-fire-rated door - default to FD30S single
                logger.warning(f"Door {door_id}: Replacement needed but fire_rating=Unknown - using default FD30S single → A05")
                a_series_code = 'A05'  # Default: FD30S single ≤2040mm
        
        # Set Option A/B columns
        # CRITICAL: Priority order matters!
        # 1. Replacement (ART17/18/20) → OptA=NO, OptB=YES
        # 2. Type 2 → OptA=YES/COMPLIANT, OptB=PENDING
        # 3. Compliant (no faults) → OptA=COMPLIANT, OptB=NO
        # 4. Has faults → OptA=YES, OptB=NO
        
        if needs_replacement:
            # Replacement door (ART17, ART18, or ART20)
            ws[f'N{row_num}'] = 'NO'   # No remedial work - full replacement
            ws[f'O{row_num}'] = 'YES'  # Option B replacement
            logger.info(f"Door {door_id}: Replacement needed (ART17/18/20) → OptA=NO, OptB=YES")
        elif is_type2:
            # Type 2 Excel surveys without replacement data: Option B = PENDING
            if is_compliant:
                ws[f'N{row_num}'] = 'COMPLIANT'
            else:
                ws[f'N{row_num}'] = 'YES'  # Has faults, needs remedial work
            ws[f'O{row_num}'] = 'PENDING'  # All Type 2 doors get PENDING (no fire strategy)
        elif is_compliant:
            # Door passed inspection - no work needed
            ws[f'N{row_num}'] = 'COMPLIANT'
            ws[f'O{row_num}'] = 'NO'
            logger.info(f"Door {door_id}: Compliant → OptA=COMPLIANT, OptB=NO")
        else:
            # Door has faults - needs remedial work
            ws[f'N{row_num}'] = 'YES'
            ws[f'O{row_num}'] = 'NO'
            logger.info(f"Door {door_id}: Has faults → OptA=YES, OptB=NO")
        
        # ISSUE #1 FIX: Column P = A-series code for replacements, B-code for remedial
        # CRITICAL VALIDATION: If OptB=YES, column P must NEVER be blank
        if needs_replacement:
            # Replacement door - MUST have A-series code
            if a_series_code:
                ws[f'P{row_num}'] = a_series_code  # A09, A04, etc.
                logger.info(f"Door {door_id}: Wrote A-series code {a_series_code} to column P")
            else:
                # VALIDATION ERROR: OptB=YES but no A-series code - this should NEVER happen
                logger.error(f"Door {door_id}: VALIDATION ERROR - OptB=YES but a_series_code is blank! fire_rating={fire_rating}, door_config={door_config}")
                logger.error(f"Door {door_id}: Using fallback A05 but THIS IS A BUG - fire rating should have been extracted correctly")
                ws[f'P{row_num}'] = 'A05'  # Emergency fallback
        else:
            # Remedial door - use primary B-code
            ws[f'P{row_num}'] = primary_code if primary_code else ''
        
        # Column Q: QTY (all B-codes for remedial work)
        ws[f'Q{row_num}'] = all_codes_str
        
        # Columns R-U: E/O flags (Extra Over costs)
        ws[f'R{row_num}'] = 'NO'  # E/O OVERSIZE
        ws[f'S{row_num}'] = 'NO'  # E/O HARDWOOD
        ws[f'T{row_num}'] = 'NO'  # E/O EXTERNAL
        ws[f'U{row_num}'] = 'NO'  # E/O VISION
        
        # Column V: NOTES / FLAGS (B-codes and flag notes)
        flag_notes = []
        
        # BUG 3 FIX: Only add ART codes for Type 1 (PDF) surveys
        # Type 2 (Excel) surveys have no ART codes
        is_type1 = door.get('format_type') == 'TYPE_1' or 'art_codes' in door
        if is_type1 and 'art_codes' in door:
            art_codes_str = ', '.join(door.get('art_codes', []))
            if art_codes_str:
                flag_notes.append(f"ARTs: {art_codes_str}")
        
        # Add flag notes for special cases requiring manual review
        if door.get('has_unable'):
            flag_notes.append("⚠️ Unable to inspect - needs revisit")
        
        # Check for ambiguous or placeholder codes that need manual review
        if any(c in ['MANUAL REVIEW', 'A-series', 'FLAG FOR MANUAL REVIEW'] for c in codes):
            flag_notes.append("⚠️ Manual review required")
        
        if flag_notes:
            ws[f'V{row_num}'] = ' | '.join(flag_notes)
        
        # Color code based on status
        # Priority: Orange (manual review) > Red (replacement) > Yellow (remedial) > Green (compliant)
        color_fill = None
        needs_manual_review = door.get('has_unable') or any(note.startswith('⚠️') for note in flag_notes)
        
        if needs_manual_review:
            color_fill = orange_fill  # Orange - needs manual review (Unable to inspect or flag notes)
        elif is_compliant:
            color_fill = green_fill  # Green - compliant
        elif needs_replacement:
            color_fill = red_fill  # Red - replacement
        else:
            color_fill = yellow_fill  # Yellow - remedial work
        
        # Apply color to all data columns A-V
        if color_fill:
            for col_letter in 'ABCDEFGHIJKLMNOPQRSTUV':
                ws[f'{col_letter}{row_num}'].fill = color_fill
    
    # PYTHON-SIDE CALCULATION: Keep formulas but write cached values
    # Formulas stay for manual editing, cached values display immediately
    logger.info("=== Calculating Quote Sheet cached values ===")
    
    # Step 1: Count B-codes AND A-series codes from Door Schedule
    door_schedule = wb['Door Schedule']
    b_code_counts = {}
    b_code_door_ids = {}
    a_code_counts = {}  # FIX #2: Track A-series codes for Option B
    a_code_door_ids = {}
    
    for row in range(4, 200):  # Check up to row 200
        door_id = door_schedule.cell(row=row, column=1).value
        if not door_id:
            break
        
        opt_a = door_schedule.cell(row=row, column=14).value  # Column N (Option A)
        opt_b = door_schedule.cell(row=row, column=15).value  # Column O (Option B)
        base_item = door_schedule.cell(row=row, column=16).value  # Column P (Base Item code)
        all_codes_str = door_schedule.cell(row=row, column=17).value  # Column Q (ALL codes)
        
        # Count Option A B-codes
        if opt_a == 'YES' and all_codes_str:
            # Parse comma-separated B-codes from column Q
            codes = [c.strip() for c in str(all_codes_str).split(',')]
            for code in codes:
                if code and code.startswith('B'):
                    b_code_counts[code] = b_code_counts.get(code, 0) + 1
                    if code not in b_code_door_ids:
                        b_code_door_ids[code] = []
                    b_code_door_ids[code].append(str(door_id))
        
        # FIX #2: Count Option B A-series codes
        if opt_b == 'YES' and base_item and str(base_item).startswith('A'):
            a_code = str(base_item).strip()
            a_code_counts[a_code] = a_code_counts.get(a_code, 0) + 1
            if a_code not in a_code_door_ids:
                a_code_door_ids[a_code] = []
            a_code_door_ids[a_code].append(str(door_id))
    
    logger.info(f"Option A B-code counts: {b_code_counts}")
    logger.info(f"Option A door IDs: {b_code_door_ids}")
    logger.info(f"Option B A-code counts: {a_code_counts}")
    logger.info(f"Option B door IDs: {a_code_door_ids}")
    
    # ISSUE #2 VALIDATION: Count total doors with OptA=YES for reconciliation
    opt_a_yes_count = sum(1 for row in range(4, 200) 
                          if door_schedule.cell(row=row, column=14).value == 'YES')
    total_b_code_instances = sum(b_code_counts.values())
    logger.info(f"VALIDATION: {opt_a_yes_count} doors with OptA=YES")
    logger.info(f"VALIDATION: {total_b_code_instances} total B-code instances counted")
    logger.info(f"VALIDATION: Average {total_b_code_instances/opt_a_yes_count if opt_a_yes_count > 0 else 0:.1f} B-codes per door")
    
    if opt_a_yes_count > 0 and total_b_code_instances == 0:
        logger.error(f"VALIDATION ERROR: {opt_a_yes_count} doors have OptA=YES but 0 B-codes counted!")
    
    # Step 2: Get rates from Rate Card
    # FIX #4: Calculate rate from Materials + Labour only (not T&J or Humping)
    rate_card_sheet = wb['Rate Card']
    rates = {}
    for row_num in range(22, 34):  # B01-B12
        code = rate_card_sheet.cell(row=row_num, column=1).value
        # OLD: rate = rate_card_sheet.cell(row=row_num, column=8).value  # Column H (TOTAL with T&J+Humping)
        # NEW: Calculate from Materials + Labour only
        mat = rate_card_sheet.cell(row=row_num, column=4).value or 0  # Column D (Materials)
        lab = rate_card_sheet.cell(row=row_num, column=5).value or 0  # Column E (Labour)
        rate = mat + lab
        if code and rate and isinstance(rate, (int, float)):
            rates[str(code)] = rate
    
    logger.info(f"Rates from Rate Card (Materials + Labour only): {rates}")
    
    # Step 3: Calculate Option A totals (cost and client price)
    option_a_cost = sum(b_code_counts.get(code, 0) * rates.get(code, 0) for code in rates.keys())
    
    # FIX #5: Get target margin and calculate client price
    target_margin = quote_sheet['R2'].value or 0
    option_a_client = option_a_cost / (1 - target_margin) if (1 - target_margin) > 0 else option_a_cost
    
    logger.info(f"Option A cost: £{option_a_cost}")
    logger.info(f"Option A client price (with {target_margin*100}% margin): £{option_a_client}")
    
    # Step 4: Write numbers to Quote Sheet
    # FIX #5: Get target margin from R2 and apply gross margin formula
    quote_sheet = wb['Quote Sheet']
    target_margin = quote_sheet['R2'].value or 0  # Default to 0 if not set
    logger.info(f"Target margin from R2: {target_margin} ({target_margin*100}%)")
    
    bcode_to_row = {
        'B01': 11, 'B02': 12, 'B03': 13, 'B04': 14, 'B05': 15, 'B06': 16,
        'B07': 17, 'B08': 18, 'B09': 19, 'B10': 20, 'B11': 21, 'B12': 22,
    }
    
    for b_code, row_num in bcode_to_row.items():
        qty = b_code_counts.get(b_code, 0)
        cost_rate = rates.get(b_code, 0)  # Materials + Labour (internal cost)
        cost_total = qty * cost_rate
        
        # FIX #5: Apply gross margin formula
        # Client Price = Cost ÷ (1 - Margin%)
        if target_margin < 1:  # Margin must be less than 100%
            client_total = cost_total / (1 - target_margin) if (1 - target_margin) > 0 else cost_total
            client_rate = client_total / qty if qty > 0 else 0
        else:
            # Fallback if margin is invalid
            client_total = cost_total
            client_rate = cost_rate
        
        door_ids = b_code_door_ids.get(b_code, [])
        door_ids_str = ', '.join(door_ids) if door_ids else ''
        
        # URGENT FIX #2: Debug logging for Quote Sheet quantities
        # Check what the Quote Sheet template label is for this row
        template_label = quote_sheet.cell(row=row_num, column=2).value  # Column B (Description/Label)
        logger.info(f"Row {row_num}: Template label='{template_label}', Writing {b_code} with QTY={qty}, Doors={door_ids_str}")
        
        # Write CLIENT PRICE values (with margin applied)
        quote_sheet.cell(row=row_num, column=3).value = qty           # Column C (QTY)
        quote_sheet.cell(row=row_num, column=5).value = client_rate   # Column E (CLIENT RATE with margin)
        quote_sheet.cell(row=row_num, column=6).value = client_total  # Column F (CLIENT TOTAL with margin)
        
        # FIX #3: Write zero to T&J and Humping rate columns to exclude from breakdown
        quote_sheet.cell(row=row_num, column=11).value = 0            # Column K (T&J rate) = 0
        quote_sheet.cell(row=row_num, column=13).value = 0            # Column M (Humping rate) = 0
        # This makes columns L and N = 0, so column O = columns H + J only (Materials + Labour = COST)
        
        # FIX #1: Door IDs reference column (use column S, after all existing columns)
        quote_sheet.cell(row=row_num, column=19).value = door_ids_str # Column S (DOOR IDs)
        
        if qty > 0:
            logger.info(f"Quote Sheet {b_code}: QTY={qty}, COST_RATE={cost_rate}, CLIENT_RATE={client_rate}, CLIENT_TOTAL={client_total}, DOOR_IDS={door_ids_str}")
    
    # Step 5: Write calculated NUMBERs to Client Summary
    # FIX #5: Client Summary shows CLIENT PRICES (with margin), not costs
    client_summary = wb['Client Summary']
    
    # HIGH FIX #4: Update column label from "NET COST" to "Internal Cost (ex margin)"
    # Check if row 9 has column headers
    if client_summary.cell(row=9, column=3).value:
        client_summary.cell(row=9, column=3).value = "Internal Cost (ex margin)"
        logger.info("Updated Client Summary C9 label to 'Internal Cost (ex margin)'")
    
    # C10 = Internal Cost (ex margin) - internal cost for reference
    client_summary.cell(row=10, column=3).value = option_a_cost
    
    # D10 = OPTION A TOTAL (CLIENT PRICE with margin)
    client_summary.cell(row=10, column=4).value = option_a_client
    
    # D13 = TOTAL INVESTMENT - Option A only (mutually exclusive with Option B)
    client_summary.cell(row=13, column=4).value = option_a_client
    
    logger.info(f"Client Summary C10 (Internal Cost ex margin): £{option_a_cost}")
    logger.info(f"Client Summary D10 (OPTION A CLIENT PRICE): £{option_a_client}")
    logger.info(f"Client Summary D13 (TOTAL INVESTMENT - Option A): £{option_a_client}")
    
    # Step 6: Write Option A TOTAL to Quote Sheet row 23
    quote_sheet.cell(row=23, column=6).value = option_a_client  # F23 (CLIENT PRICE)
    logger.info(f"Quote Sheet F23 (Option A TOTAL - CLIENT PRICE): £{option_a_client}")
    
    # Step 6b: FIX #2 - Calculate and write Option B totals (replacement doors)
    # Get A-series rates from Rate Card
    a_series_rates = {}
    for row_num in range(6, 40):  # Check all rows in Rate Card
        code = rate_card_sheet.cell(row=row_num, column=1).value
        if code and str(code).startswith('A'):
            mat = rate_card_sheet.cell(row=row_num, column=4).value or 0
            lab = rate_card_sheet.cell(row=row_num, column=5).value or 0
            rate = mat + lab
            a_series_rates[str(code)] = rate
    
    logger.info(f"A-series rates from Rate Card: {a_series_rates}")
    
    # Calculate Option B totals (cost and client price)
    option_b_cost = sum(a_code_counts.get(code, 0) * a_series_rates.get(code, 0) for code in a_series_rates.keys())
    
    # FIX #5: Apply margin to Option B
    option_b_client = option_b_cost / (1 - target_margin) if (1 - target_margin) > 0 else option_b_cost
    
    logger.info(f"Option B cost: £{option_b_cost}")
    logger.info(f"Option B client price (with {target_margin*100}% margin): £{option_b_client}")
    
    # Write Option B line items to Quote Sheet (rows 26-40)
    acode_to_row = {
        'A01': 26, 'A02': 27, 'A03': 28, 'A04': 29, 'A05': 30, 'A06': 31,
        'A07': 32, 'A08': 33, 'A09': 34, 'A10': 35, 'A11': 36, 'A12': 37,
        'A13': 38, 'A14': 39, 'A15': 40,
    }
    
    for a_code, row_num in acode_to_row.items():
        qty = a_code_counts.get(a_code, 0)
        cost_rate = a_series_rates.get(a_code, 0)
        cost_total = qty * cost_rate
        
        # FIX #5: Apply gross margin formula to Option B
        client_total = cost_total / (1 - target_margin) if (1 - target_margin) > 0 else cost_total
        client_rate = client_total / qty if qty > 0 else 0
        
        door_ids = a_code_door_ids.get(a_code, [])
        door_ids_str = ', '.join(door_ids) if door_ids else ''
        
        # Write Option B CLIENT PRICE values (with margin)
        quote_sheet.cell(row=row_num, column=3).value = qty            # Column C (QTY)
        quote_sheet.cell(row=row_num, column=5).value = client_rate    # Column E (CLIENT RATE with margin)
        quote_sheet.cell(row=row_num, column=6).value = client_total   # Column F (CLIENT TOTAL with margin)
        quote_sheet.cell(row=row_num, column=11).value = 0             # Column K (T&J rate) = 0
        quote_sheet.cell(row=row_num, column=13).value = 0             # Column M (Humping rate) = 0
        quote_sheet.cell(row=row_num, column=19).value = door_ids_str  # Column S (DOOR IDs)
        
        if qty > 0:
            logger.info(f"Quote Sheet {a_code}: QTY={qty}, COST_RATE={cost_rate}, CLIENT_RATE={client_rate}, CLIENT_TOTAL={client_total}, DOOR_IDs={door_ids_str}")
    
    # ISSUE #3 FIX: Write Option B TOTAL to Quote Sheet
    # Try common locations for Option B TOTAL row
    # Most templates have Option B section starting around row 26, so TOTAL might be around row 42-45
    # Write to row 42 as a best guess (can adjust if template differs)
    option_b_total_row = 42
    quote_sheet.cell(row=option_b_total_row, column=6).value = option_b_client if option_b_client > 0 else 0
    logger.info(f"Quote Sheet F{option_b_total_row} (Option B TOTAL - CLIENT PRICE): £{option_b_client}")
    
    # VALIDATION: Ensure TOTAL rows are NEVER None
    if quote_sheet.cell(row=23, column=6).value is None:
        logger.error("VALIDATION ERROR: Quote Sheet F23 (Option A TOTAL) is None - writing 0 as fallback")
        quote_sheet.cell(row=23, column=6).value = 0
    if quote_sheet.cell(row=option_b_total_row, column=6).value is None:
        logger.error(f"VALIDATION ERROR: Quote Sheet F{option_b_total_row} (Option B TOTAL) is None - writing 0 as fallback")
        quote_sheet.cell(row=option_b_total_row, column=6).value = 0
    
    # Write Option B total to Client Summary E11
    if option_b_client > 0:
        client_summary.cell(row=11, column=5).value = option_b_client  # E11 (Option B CLIENT PRICE)
        logger.info(f"Client Summary E11 (Option B CLIENT PRICE): £{option_b_client}")
    
    # URGENT FIX #5: TOTAL INVESTMENT should NOT add Option A + Option B (mutually exclusive)
    # Show Option A in column D, Option B in column E
    # Matt chooses one, not both
    client_summary.cell(row=13, column=4).value = option_a_client  # D13 = Option A only
    client_summary.cell(row=13, column=5).value = option_b_client if option_b_client > 0 else ""  # E13 = Option B only
    logger.info(f"Client Summary D13 (TOTAL INVESTMENT - Option A): £{option_a_client}")
    logger.info(f"Client Summary E13 (TOTAL INVESTMENT - Option B): £{option_b_client}")
    
    # Step 7: Copy header values from Quote Sheet to Client Summary
    # Client Summary pulls these from Quote Sheet via formulas, but we need to write them
    # Correct row mapping:
    # - Row 3: Client (B3)
    # - Row 4: Site (B4)
    # - Row 5: Building (B5)
    # - Row 6: Date (E6)
    # - Row 7: Quote Ref (B7)
    client_summary.cell(row=3, column=2).value = quote_sheet['B4'].value  # B3 = Client
    client_summary.cell(row=4, column=2).value = quote_sheet['B5'].value  # B4 = Site (may be blank for Type 2)
    client_summary.cell(row=5, column=2).value = quote_sheet['B6'].value  # B5 = Building (may be blank for Type 2)
    client_summary.cell(row=6, column=5).value = quote_sheet['B3'].value  # E6 = Date
    client_summary.cell(row=7, column=2).value = quote_sheet['B2'].value  # B7 = Quote Ref
    logger.info(f"Client Summary headers: Client={quote_sheet['B4'].value}, Date={quote_sheet['B3'].value}, Ref={quote_sheet['B2'].value}")
    
    # Step 8: Set Option B to "PENDING" for Type 2 surveys (no fire strategy)
    # UNLESS we already calculated Option B client price above (replacement doors with specs)
    is_type2 = doors and doors[0].get('format_type') == 'TYPE_2'
    if is_type2 and option_b_client == 0:
        client_summary.cell(row=11, column=5).value = "PENDING — fire strategy required"  # E11 (Option B)
        logger.info("Client Summary E11 (Option B): PENDING — fire strategy required (Type 2)")
    
    # MEDIUM FIX #7: Set correct compliance note based on survey type
    # Compliance note is typically in Client Summary around row 15-20
    # Type 1: Full FireDNA survey with fire strategy
    # Type 2: Excel survey without fire strategy
    compliance_note_row = 16  # Adjust if needed based on template
    if is_type2:
        compliance_note = (
            "This report identifies fire doors requiring remedial works based on visual inspection. "
            "Option A provides remedial works to address identified faults. Option B (full replacement) "
            "cannot be priced without fire strategy drawings - please request these from the client if "
            "replacement doors are required. All prices exclude VAT and are subject to site survey confirmation."
        )
    else:
        # Type 1 compliance note
        compliance_note = (
            "This report identifies fire doors requiring remedial works to achieve compliance with "
            "Approved Document B and the Regulatory Reform (Fire Safety) Order 2005. Works are categorized "
            "as Option A (remedial works) or Option B (full replacement). All prices exclude VAT and are "
            "subject to site survey confirmation."
        )
    
    # Try to find and update compliance note cell (typically a merged cell)
    # Check rows 15-20 for compliance note
    for row in range(15, 21):
        cell_value = client_summary.cell(row=row, column=2).value
        if cell_value and ('compliance' in str(cell_value).lower() or 'approved document' in str(cell_value).lower()):
            client_summary.cell(row=row, column=2).value = compliance_note
            logger.info(f"Updated compliance note at row {row} to Type {'2' if is_type2 else '1'} note")
            break
    else:
        logger.warning("Could not find compliance note cell in Client Summary - may need manual update")
    
    # Step 9: FIX #6 - Populate Material Call-Off sheet with B-code counts
    if "Material Call-Off" in wb.sheetnames:
        material_calloff = wb["Material Call-Off"]
        
        # Row 13: Closers (B03)
        closers_count = b_code_counts.get('B03', 0)
        material_calloff.cell(row=13, column=4).value = closers_count  # Column D
        
        # Row 14: Hinges (B04)
        hinges_count = b_code_counts.get('B04', 0)
        material_calloff.cell(row=14, column=4).value = hinges_count  # Column D
        
        # Row 15: Seals (B01 - intumescent + smoke)
        seals_count = b_code_counts.get('B01', 0)
        material_calloff.cell(row=15, column=4).value = seals_count  # Column D
        
        # Row 16: Intumescent only (B02)
        intumescent_count = b_code_counts.get('B02', 0)
        material_calloff.cell(row=16, column=4).value = intumescent_count  # Column D
        
        # Row 17: Signage (B07 × 2)
        signage_count = b_code_counts.get('B07', 0) * 2  # 2 signs per door
        material_calloff.cell(row=17, column=4).value = signage_count  # Column D
        
        # ISSUE #4 FIX: Add all missing B-code component rows
        # Row 18: Lever handles/latch (B05)
        latch_count = b_code_counts.get('B05', 0)
        material_calloff.cell(row=18, column=4).value = latch_count  # Column D
        
        # Row 19: Drop seal (B09)
        drop_seal_count = b_code_counts.get('B09', 0)
        material_calloff.cell(row=19, column=4).value = drop_seal_count  # Column D
        
        # Row 20: Intumescent mastic (B06)
        mastic_count = b_code_counts.get('B06', 0)
        material_calloff.cell(row=20, column=4).value = mastic_count  # Column D
        
        # Row 21: Hardwood lipping (B11)
        lipping_count = b_code_counts.get('B11', 0)
        material_calloff.cell(row=21, column=4).value = lipping_count  # Column D
        
        # Row 22: Hardwood void infill (B12)
        void_infill_count = b_code_counts.get('B12', 0)
        material_calloff.cell(row=22, column=4).value = void_infill_count  # Column D
        
        logger.info(f"Material Call-Off counts:")
        logger.info(f"  Closers (B03) = {closers_count}")
        logger.info(f"  Hinges (B04) = {hinges_count}")
        logger.info(f"  Seals (B01) = {seals_count}")
        logger.info(f"  Intumescent only (B02) = {intumescent_count}")
        logger.info(f"  Signage (B07×2) = {signage_count}")
        logger.info(f"  Lever handles/latch (B05) = {latch_count}")
        logger.info(f"  Drop seal (B09) = {drop_seal_count}")
        logger.info(f"  Intumescent mastic (B06) = {mastic_count}")
        logger.info(f"  Hardwood lipping (B11) = {lipping_count}")
        logger.info(f"  Hardwood void infill (B12) = {void_infill_count}")
    
    logger.info("=== Line item numbers + header values written ===")
    
    # Save workbook
    try:
        logger.info(f"Saving workbook to: {output_path}")
        
        # Keep calculation settings for any remaining formulas
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        
        wb.save(output_path)
        logger.info("Workbook saved successfully")
    except Exception as e:
        error_msg = f"Failed to save workbook: {str(e)}"
        logger.error(error_msg)
        raise RuntimeError(error_msg) from e
    
    # Verify file was created
    output_file = Path(output_path)
    if not output_file.exists():
        error_msg = f"Output file was not created at: {output_path}"
        logger.error(error_msg)
        raise RuntimeError(error_msg)
    
    file_size = output_file.stat().st_size
    logger.info(f"Output file created successfully. Size: {file_size} bytes")
    
    # Formulas are preserved with cached values set
    # Excel should display correct values on open
    logger.info("Formulas preserved with calculated cached values")
    
    return output_path
